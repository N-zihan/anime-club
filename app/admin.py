"""
南平一中动漫社官网 · 管理后台模块
====================================

本模块为社团管理层提供完整的后台管理功能，包含以下子系统：

1. 活动管理 —— 增删改社团活动，支持活动关联照片的级联删除
2. 照片墙管理 —— 上传/删除活动照片，存储至 Supabase Storage
3. 番剧资源管理 —— 审核社员推荐的番剧，或手动添加资源
4. 用户管理 —— 查看社员列表，切换运营/站长身份，删除用户

权限体系：
- admin_required 装饰器：允许站长或运营访问
- owner_required 装饰器：仅允许站长访问（用于敏感操作）

后台入口：
- /admin/entry 统一入口，根据角色自动分流至站长或运营面板
"""

import io
import uuid
from datetime import datetime, timedelta
from functools import wraps

from flask import Blueprint, render_template, request, redirect, url_for, session, flash
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .models import db, User, Activity, Photo, AnimeResource, Message, Reply, Contest, Nomination, Candidate, \
    ContestVote
from .utils import get_supabase, allowed_file, compress_image, get_or_404

admin_bp = Blueprint('admin', __name__)
supabase = get_supabase()


# ---------- 权限装饰器 ----------
def admin_required(f):
    """运营或站长均可访问，直接从 Session 读取角色"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        role = session.get('user_role')
        if role not in ('owner', 'staff'):
            flash('你没有管理权限', 'danger')
            return redirect(url_for('public.index'))
        return f(*args, **kwargs)

    return decorated_function


def owner_required(f):
    """仅站长可访问，直接从 Session 读取角色"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('user_role') != 'owner':
            flash('该功能仅站长可用', 'danger')
            return redirect(url_for('admin.dashboard'))
        return f(*args, **kwargs)

    return decorated_function


# ---------- 统一入口 ----------
@admin_bp.route('/admin/entry')
def admin_entry():
    role = session.get('user_role')
    if role == 'owner':
        return redirect(url_for('admin.dashboard'))
    elif role == 'staff':
        return redirect(url_for('admin.staff_dashboard'))
    else:
        flash('你没有管理权限', 'danger')
        return redirect(url_for('public.index'))


# ---------- 站长后台 ----------
@admin_bp.route('/admin/dashboard')
@admin_required
def dashboard():
    return render_template('admin_dashboard.html')


# ---------- 运营后台 ----------
@admin_bp.route('/staff/dashboard')
@admin_required
def staff_dashboard():
    return render_template('staff_dashboard.html')


# ---------- 活动管理 ----------
@admin_bp.route('/admin/activities')
@admin_required
def admin_activities():
    activities = Activity.query.order_by(Activity.date.desc()).all()
    return render_template('admin_activities.html', activities=activities)


@admin_bp.route('/admin/activities/add', methods=['GET', 'POST'])
@admin_required
def admin_activity_add():
    if request.method == 'POST':
        title = request.form['title']
        date = request.form['date']
        content = request.form['content']
        new_activity = Activity(title=title, date=date, content=content)
        db.session.add(new_activity)
        db.session.commit()
        flash('活动添加成功', 'success')
        return redirect(url_for('admin.admin_activities'))
    return render_template('admin_activity_form.html', activity=None)


@admin_bp.route('/admin/activities/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def admin_activity_edit(id):
    activity = get_or_404(Activity, id)
    if request.method == 'POST':
        activity.title = request.form['title']
        activity.date = request.form['date']
        activity.content = request.form['content']
        db.session.commit()
        flash('活动已更新', 'success')
        return redirect(url_for('admin.admin_activities'))
    return render_template('admin_activity_form.html', activity=activity)


@admin_bp.route('/admin/activities/delete/<int:id>')
@admin_required
def admin_activity_delete(id):
    activity = get_or_404(Activity, id)

    # 先删除该活动关联的所有照片（文件 + 数据库记录）
    for photo in activity.photos:
        try:
            supabase.storage.from_('photos').remove([photo.filename])
        except Exception:
            pass  # 云文件删不掉也继续，至少删数据库记录
        db.session.delete(photo)

    db.session.delete(activity)
    db.session.commit()
    flash('活动及其所有照片已删除', 'success')
    return redirect(url_for('admin.admin_activities'))


# ---------- 照片墙管理 ----------
@admin_bp.route('/admin/gallery')
@admin_required
def admin_gallery():
    activities = Activity.query.order_by(Activity.date.desc()).all()
    photos = Photo.query.order_by(Photo.upload_time.desc()).all()
    for photo in photos:
        photo.url = supabase.storage.from_('photos').get_public_url(photo.filename)
    return render_template('admin_gallery.html', activities=activities, photos=photos)


@admin_bp.route('/admin/gallery/upload', methods=['POST'])
@admin_required
def admin_gallery_upload():
    if not session.get('user_id'):
        flash('请先登录再上传照片', 'warning')
        return redirect(url_for('auth.login'))

    if 'file' not in request.files:
        flash('没有文件', 'danger')
        return redirect(url_for('admin.admin_gallery'))
    file = request.files['file']
    if file.filename == '':
        flash('未选择文件', 'danger')
        return redirect(url_for('admin.admin_gallery'))

    activity_id = request.form.get('activity_id')
    if activity_id:
        activity_id = int(activity_id)

    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.{ext}"

        try:
            raw_data = file.read()
            # 照片墙压缩到 1200x1200，品质85
            compressed = compress_image(raw_data, max_size=(1200, 1200), quality=85)
            supabase.storage.from_('photos').upload(
                filename,
                compressed,
                file_options={"content-type": 'image/jpeg'}
            )
        except Exception as e:
            flash(f'上传失败: {str(e)}', 'danger')
            return redirect(url_for('admin.admin_gallery'))

        photo = Photo(
            filename=filename,
            activity_id=activity_id if activity_id else None,
            uploader=session.get('username', '社员')  # 记录上传者
        )
        db.session.add(photo)
        db.session.commit()
        flash('照片上传成功', 'success')
    else:
        flash('不支持的文件类型', 'danger')
    return redirect(url_for('admin.admin_gallery'))


@admin_bp.route('/admin/gallery/delete/<int:photo_id>')
@admin_required
def admin_gallery_delete(photo_id):
    """删除单张照片（文件 + 数据库记录）"""
    photo = get_or_404(Photo, photo_id)

    # 从 Supabase Storage 删除文件
    try:
        supabase.storage.from_('photos').remove([photo.filename])
    except Exception as e:
        # 云文件删不掉也继续，至少删数据库记录
        flash(f'云存储文件删除失败（但数据库记录已删）: {str(e)}', 'warning')

    db.session.delete(photo)
    db.session.commit()

    flash('照片已删除', 'success')
    return redirect(url_for('admin.admin_gallery'))


# ---------- 番剧资源管理 ----------
@admin_bp.route('/admin/anime_resources')
@admin_required
def admin_anime_resources():
    pending = AnimeResource.query.filter_by(status='pending').order_by(AnimeResource.upload_time.desc()).all()
    approved = AnimeResource.query.filter_by(status='approved').order_by(AnimeResource.upload_time.desc()).all()
    return render_template('admin_anime_resources.html', pending_resources=pending, approved_resources=approved)


@admin_bp.route('/admin/anime_resources/approve/<int:id>')
@admin_required
def approve_anime_resource(id):
    resource = get_or_404(AnimeResource, id)
    resource.status = 'approved'
    db.session.commit()
    flash('已通过审核', 'success')
    return redirect(url_for('admin.admin_anime_resources'))


@admin_bp.route('/admin/anime_resources/reject/<int:id>')
@admin_required
def reject_anime_resource(id):
    resource = get_or_404(AnimeResource, id)
    db.session.delete(resource)
    db.session.commit()
    flash('已拒绝并删除', 'warning')
    return redirect(url_for('admin.admin_anime_resources'))


@admin_bp.route('/admin/anime_resources/delete/<int:id>')
@admin_required
def admin_anime_resources_delete(id):
    resource = get_or_404(AnimeResource, id)
    db.session.delete(resource)
    db.session.commit()
    flash('已删除', 'success')
    return redirect(url_for('admin.admin_anime_resources'))


@admin_bp.route('/admin/anime_resources/add', methods=['GET', 'POST'])
@admin_required
def admin_anime_resources_add():
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        link = request.form.get('link')
        extract_code = request.form.get('extract_code')
        if not title or not link:
            flash('标题和链接不能为空', 'danger')
            return redirect(url_for('admin.admin_anime_resources_add'))
        resource = AnimeResource(
            title=title,
            description=description,
            link=link,
            extract_code=extract_code,
            user_id=session.get('user_id'),  # 管理员添加时，以当前登录用户为提交人
            status='approved'
        )
        db.session.add(resource)
        db.session.commit()
        flash('资源添加成功', 'success')
        return redirect(url_for('admin.admin_anime_resources'))
    return render_template('admin_anime_resources_add.html')


# ---------- 用户管理 ----------
@admin_bp.route('/admin/users')
@admin_required
@owner_required
def admin_users():
    users = User.query.order_by(User.registered_at.desc()).all()
    return render_template('admin_users.html', users=users)


@admin_bp.route('/admin/users/delete/<int:user_id>')
@admin_required
def admin_delete_user(user_id):
    user = get_or_404(User, user_id)
    if user.id == session.get('user_id'):
        flash('不能删除自己', 'danger')
        return redirect(url_for('admin.admin_users'))

    # 级联删除所有关联数据
    # 注意：删除顺序很重要，先删子表，再删主表

    # 1. 删除该用户的提名（Nomination）
    Nomination.query.filter_by(user_id=user.id).delete()

    # 2. 删除该用户的投票（ContestVote）
    ContestVote.query.filter_by(user_id=user.id).delete()

    # 3. 删除该用户的留言（Message）和回复（Reply）
    # 留言：先删回复，再删留言
    for msg in Message.query.filter_by(user_id=user.id).all():
        Reply.query.filter_by(message_id=msg.id).delete()
    Message.query.filter_by(user_id=user.id).delete()

    # 4. 删除该用户的番剧推荐（AnimeResource）
    AnimeResource.query.filter_by(user_id=user.id).delete()

    # 5. 最后删除用户
    db.session.delete(user)
    db.session.commit()

    flash('用户及其所有关联数据已删除', 'success')
    return redirect(url_for('admin.admin_users'))


@admin_bp.route('/admin/users/toggle_staff/<int:user_id>')
@admin_required
def admin_toggle_staff(user_id):
    user = get_or_404(User, user_id)
    user.is_staff = not user.is_staff
    db.session.commit()
    flash(f'用户 {user.username} 的运营状态已更新', 'success')
    return redirect(url_for('admin.admin_users'))


@admin_bp.route('/admin/users/toggle_owner/<int:user_id>')
@admin_required
def admin_toggle_owner(user_id):
    user = get_or_404(User, user_id)
    if not user.is_owner:
        User.query.update({User.is_owner: False})
        db.session.commit()
    user.is_owner = not user.is_owner
    db.session.commit()
    flash(f'用户 {user.username} 的站长状态已更新', 'success')
    return redirect(url_for('admin.admin_users'))


@admin_bp.route('/admin/users/export')
@admin_required
def export_users_excel():
    users = User.query.order_by(User.registered_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '社员名单'

    # 表头
    headers = ['ID', '用户名', 'QQ号', '注册时间', '运营', '站长']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=user.id)
        ws.cell(row=row_idx, column=2, value=user.username)
        ws.cell(row=row_idx, column=3, value=user.qq)
        ws.cell(row=row_idx, column=4, value=user.registered_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_idx, column=5, value='是' if user.is_staff else '否')
        ws.cell(row=row_idx, column=6, value='是' if user.is_owner else '否')

    # 列宽自适应
    for col in range(1, 7):
        max_length = 0
        for row in range(1, len(users) + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[chr(64 + col)].width = min(max_length + 2, 30)  # A=65, B=66, ...

    # 写入内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'社员名单_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ---------- 留言管理 ----------
@admin_bp.route('/admin/messages')
@admin_required
def admin_messages():
    """留言管理页面，显示所有留言及回复"""
    messages = Message.query.order_by(Message.timestamp.desc()).all()
    return render_template('admin_messages.html', messages=messages)


@admin_bp.route('/admin/messages/delete/<int:message_id>')
@admin_required
def admin_delete_message(message_id):
    """删除留言及其所有回复"""
    message = get_or_404(Message, message_id)
    # 删除所有回复（如果有外键级联，可以直接删除message，但为了明确，手动删除回复）
    for reply in message.replies.all():
        db.session.delete(reply)
    db.session.delete(message)
    db.session.commit()
    flash('留言及其所有回复已删除', 'success')
    return redirect(url_for('admin.admin_messages'))


@admin_bp.route('/admin/replies/delete/<int:reply_id>')
@admin_required
def admin_delete_reply(reply_id):
    """删除单条回复，保留留言和其他回复"""
    reply = get_or_404(Reply, reply_id)
    db.session.delete(reply)
    db.session.commit()
    flash('回复已删除', 'success')
    return redirect(url_for('admin.admin_messages'))


# ========== 萌战系统 · 管理后台 ==========

@admin_bp.route('/admin/contests/manage')
@admin_required
def admin_contests_manage():
    contests = Contest.query.order_by(Contest.created_at.desc()).all()
    return render_template('admin_contests_manage.html', contests=contests)


@admin_bp.route('/admin/contests/create', methods=['GET', 'POST'])
@admin_required
def admin_contest_create():
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        gender_mode = request.form.get('gender_mode', 'separate')
        open_at_str = request.form.get('open_at')

        if not title or not description or not open_at_str:
            flash('标题、描述和开始时间不能为空', 'danger')
            return redirect(url_for('admin.admin_contest_create'))

        open_at = datetime.strptime(open_at_str, '%Y-%m-%dT%H:%M')

        contest = Contest(
            title=title,
            description=description,
            type='saimoe',
            gender_mode=gender_mode,
            status='draft',
            created_by=session.get('user_id'),
            open_at=open_at,
            close_at=open_at + timedelta(days=50),  # 自动计算结束时间
        )
        db.session.add(contest)
        db.session.commit()
        flash(f'赛事 "{title}" 创建成功', 'success')
        return redirect(url_for('admin.admin_contest_edit', contest_id=contest.id))

    return render_template('admin_contest_create.html')


@admin_bp.route('/admin/contests/edit/<int:contest_id>', methods=['GET', 'POST'])
@admin_required
def admin_contest_edit(contest_id):
    contest = get_or_404(Contest, contest_id)

    if request.method == 'POST':
        contest.title = request.form.get('title')
        contest.description = request.form.get('description')
        contest.gender_mode = request.form.get('gender_mode', 'separate')
        contest.status = request.form.get('status')

        open_at_str = request.form.get('open_at')
        if open_at_str:
            contest.open_at = datetime.strptime(open_at_str, '%Y-%m-%dT%H:%M')
            contest.close_at = contest.open_at + timedelta(days=50)  # 更新开始时间时自动刷新结束时间

        db.session.commit()
        flash('赛事信息已更新', 'success')
        return redirect(url_for('admin.admin_contest_edit', contest_id=contest.id))

    pending_nominations = contest.nominations.filter_by(status='pending').all()
    candidates = contest.candidates.all()
    return render_template('admin_contest_edit.html',
                           contest=contest,
                           nominations=pending_nominations,
                           candidates=candidates)


@admin_bp.route('/admin/contests/delete/<int:contest_id>')
@admin_required
def admin_contest_delete(contest_id):
    contest = get_or_404(Contest, contest_id)

    # 1. 手动删除所有关联的候选角色 (Candidates)
    #    使用 delete() 直接执行 SQL，比循环删除更高效
    Candidate.query.filter_by(contest_id=contest.id).delete()

    # 2. 手动删除所有关联的提名 (Nominations)
    Nomination.query.filter_by(contest_id=contest.id).delete()

    # 3. 最后删除赛事本身
    db.session.delete(contest)
    db.session.commit()

    flash('赛事已删除', 'success')
    return redirect(url_for('admin.admin_contests_manage'))


# ---------- 提名审核 ----------
@admin_bp.route('/admin/nominations/approve/<int:nomination_id>')
@admin_required
def admin_nomination_approve(nomination_id):
    nomination = get_or_404(Nomination, nomination_id)
    contest = nomination.contest

    existing = Candidate.query.filter_by(
        contest_id=contest.id,
        name=nomination.name
    ).first()
    if existing:
        flash(f'角色 "{nomination.name}" 已存在于候选池', 'danger')
        nomination.status = 'rejected'
        db.session.commit()
        return redirect(url_for('admin.admin_contest_edit', contest_id=contest.id))

    candidate = Candidate(
        contest_id=contest.id,
        nomination_id=nomination.id,
        name=nomination.name,
        source=nomination.source,
        gender=nomination.gender,
        image_url=nomination.image_url,
        description=nomination.description,
        stage='pending'  # 初始状态
    )
    db.session.add(candidate)
    nomination.status = 'approved'
    db.session.commit()
    flash(f'已通过提名: {nomination.name}', 'success')
    return redirect(url_for('admin.admin_contest_edit', contest_id=contest.id))


@admin_bp.route('/admin/nominations/reject/<int:nomination_id>')
@admin_required
def admin_nomination_reject(nomination_id):
    nomination = get_or_404(Nomination, nomination_id)
    contest_id = nomination.contest_id
    nomination.status = 'rejected'
    db.session.commit()
    flash(f'已拒绝提名: {nomination.name}', 'warning')
    return redirect(url_for('admin.admin_contest_edit', contest_id=contest_id))


@admin_bp.route('/admin/candidates/delete/<int:candidate_id>')
@admin_required
def admin_candidate_delete(candidate_id):
    candidate = get_or_404(Candidate, candidate_id)
    contest_id = candidate.contest_id
    db.session.delete(candidate)
    db.session.commit()
    flash('已从候选池移除该角色', 'success')
    return redirect(url_for('admin.admin_contest_edit', contest_id=contest_id))
